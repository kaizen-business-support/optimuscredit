"""
Tests unitaires pour le module excel_loader.py
"""

import unittest
import sys
import os
import tempfile
import openpyxl
from unittest.mock import patch, MagicMock
from pathlib import Path

# Ajouter le dossier parent au path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from modules.core.excel_loader import ExcelDataLoader


class TestExcelDataLoader(unittest.TestCase):
    """Tests pour la classe ExcelDataLoader"""
    
    def setUp(self):
        """Configuration initiale des tests"""
        self.loader = ExcelDataLoader()
        
        # Créer un fichier Excel temporaire pour les tests
        self.temp_dir = tempfile.mkdtemp()
        self.temp_excel_path = os.path.join(self.temp_dir, "test_financial_data.xlsx")
        self.create_test_excel_file()
    
    def tearDown(self):
        """Nettoyage après les tests"""
        # Supprimer les fichiers temporaires
        if os.path.exists(self.temp_excel_path):
            os.remove(self.temp_excel_path)
        os.rmdir(self.temp_dir)
    
    def create_test_excel_file(self):
        """Crée un fichier Excel de test avec la structure BCEAO"""
        workbook = openpyxl.Workbook()
        
        # Supprimer la feuille par défaut
        workbook.remove(workbook.active)
        
        # === FEUILLE BILAN ===
        bilan_sheet = workbook.create_sheet(title="Bilan")
        
        # Structure du bilan avec données de test
        bilan_data = {
            # Actif
            'E11': 200000,  # Terrains
            'E12': 500000,  # Bâtiments
            'E13': 50000,   # Agencements
            'E14': 300000,  # Matériel et mobilier
            'E15': 100000,  # Matériel de transport
            'E16': 25000,   # Avances immobilisations
            'E19': 75000,   # Titres participation
            'E20': 50000,   # Autres immob financières
            'E21': 1300000, # Total immobilisations nettes
            
            'E22': 0,       # Actif circulant HAO
            'E23': 150000,  # Stocks
            'E25': 20000,   # Fournisseurs avances versées
            'E26': 100000,  # Créances clients
            'E27': 50000,   # Autres créances
            'E28': 320000,  # Total actif circulant
            
            'E30': 25000,   # Titres placement
            'E31': 5000,    # Valeurs à encaisser
            'E32': 50000,   # Banques caisses
            'E33': 80000,   # Total trésorerie actif
            'E35': 1700000, # Total général actif
            
            # Passif
            'I5': 200000,   # Capital
            'I6': 0,        # Capital non appelé
            'I7': 50000,    # Primes capital
            'I8': 0,        # Écarts réévaluation
            'I9': 75000,    # Réserves indisponibles
            'I10': 125000,  # Réserves libres
            'I11': 0,       # Report à nouveau
            'I12': 150000,  # Résultat net
            'I13': 0,       # Subventions investissement
            'I14': 0,       # Provisions réglementées
            'I15': 600000,  # Total capitaux propres
            
            'I17': 400000,  # Emprunts dettes financières
            'I18': 0,       # Dettes location acquisition
            'I19': 50000,   # Provisions financières
            'I20': 450000,  # Total dettes financières
            
            'I22': 0,       # Dettes circulantes HAO
            'I23': 10000,   # Clients avances reçues
            'I24': 120000,  # Fournisseurs exploitation
            'I25': 80000,   # Dettes sociales fiscales
            'I26': 40000,   # Autres dettes
            'I27': 0,       # Provisions risques CT
            'I28': 250000,  # Total passif circulant
            
            'I30': 0,       # Banques crédits escompte
            'I31': 50000,   # Banques crédits trésorerie
            'I33': 50000,   # Total trésorerie passif
            'I34': 0,       # Écart conversion passif
        }
        
        for cell_ref, value in bilan_data.items():
            bilan_sheet[cell_ref] = value
        
        # === FEUILLE COMPTE DE RÉSULTAT ===
        cr_sheet = workbook.create_sheet(title="CR")
        
        cr_data = {
            'E5': 200000,   # Ventes marchandises
            'E6': 120000,   # Achats marchandises
            'E7': 0,        # Variation stocks marchandises
            'E8': 80000,    # Marge commerciale
            
            'E9': 1200000,  # Ventes produits fabriqués
            'E10': 300000,  # Travaux services vendus
            'E11': 0,       # Produits accessoires
            'E12': 1700000, # Chiffre d'affaires
            
            'E13': 0,       # Production stockée
            'E14': 0,       # Production immobilisée
            'E15': 0,       # Subventions exploitation
            'E16': 25000,   # Autres produits
            'E17': 0,       # Transferts charges exploitation
            
            'E18': 600000,  # Achats matières premières
            'E19': 0,       # Variation stocks MP
            'E20': 200000,  # Autres achats
            'E21': 0,       # Variation stocks autres
            'E22': 50000,   # Transports
            'E23': 100000,  # Services extérieurs
            'E24': 30000,   # Impôts taxes
            'E25': 40000,   # Autres charges
            
            'E26': 785000,  # Valeur ajoutée
            'E27': 350000,  # Charges personnel
            'E28': 435000,  # Excédent brut
            
            'E29': 0,       # Reprises amortissements
            'E30': 100000,  # Dotations amortissements
            'E31': 335000,  # Résultat exploitation
            
            'E32': 5000,    # Revenus financiers
            'E33': 0,       # Reprises provisions financières
            'E34': 0,       # Transferts charges financières
            'E35': 25000,   # Frais financiers
            'E36': 0,       # Dotations provisions financières
            'E37': -20000,  # Résultat financier
            
            'E38': 315000,  # Résultat activités ordinaires
            
            'E39': 0,       # Produits cessions immob
            'E40': 0,       # Autres produits HAO
            'E41': 0,       # Valeurs comptables cessions
            'E42': 0,       # Autres charges HAO
            'E43': 0,       # Résultat HAO
            'E44': 0,       # Participation travailleurs
            'E45': 15000,   # Impôts résultat
            'E46': 150000,  # Résultat net
        }
        
        for cell_ref, value in cr_data.items():
            cr_sheet[cell_ref] = value
        
        # === FEUILLE TABLEAU DES FLUX ===
        tft_sheet = workbook.create_sheet(title="TFT")
        
        tft_data = {
            'E3': 30000,    # Trésorerie ouverture
            'E5': 250000,   # CAFG
            'E11': 200000,  # Flux activités opérationnelles
            'E18': -150000, # Flux activités investissement
            'E24': 0,       # Flux capitaux propres
            'E29': -20000,  # Flux capitaux étrangers
            'E30': -20000,  # Flux activités financement
            'E31': 30000,   # Variation trésorerie
            'E32': 80000,   # Trésorerie clôture
        }
        
        for cell_ref, value in tft_data.items():
            tft_sheet[cell_ref] = value
        
        # Sauvegarder le fichier
        workbook.save(self.temp_excel_path)
        workbook.close()
    
    def create_invalid_excel_file(self, filename="invalid.xlsx"):
        """Crée un fichier Excel invalide pour tester la gestion d'erreurs"""
        invalid_path = os.path.join(self.temp_dir, filename)
        workbook = openpyxl.Workbook()
        
        # Créer une feuille avec un nom incorrect
        sheet = workbook.active
        sheet.title = "WrongSheetName"
        sheet['A1'] = "Invalid data"
        
        workbook.save(invalid_path)
        workbook.close()
        return invalid_path
    
    def test_load_excel_template_success(self):
        """Test du chargement réussi d'un fichier Excel conforme"""
        data = self.loader.load_excel_template(self.temp_excel_path)
        
        # Vérifier que les données sont extraites
        self.assertIsNotNone(data)
        self.assertIsInstance(data, dict)
        
        # Vérifier quelques valeurs clés
        self.assertEqual(data.get('total_actif'), 1700000)
        self.assertEqual(data.get('capitaux_propres'), 600000)
        self.assertEqual(data.get('chiffre_affaires'), 1700000)
        self.assertEqual(data.get('resultat_net'), 150000)
        
        # Vérifier des valeurs du bilan
        self.assertEqual(data.get('terrains'), 200000)
        self.assertEqual(data.get('batiments'), 500000)
        self.assertEqual(data.get('stocks'), 150000)
        self.assertEqual(data.get('creances_clients'), 100000)
        
        # Vérifier des valeurs du compte de résultat
        self.assertEqual(data.get('valeur_ajoutee'), 785000)
        self.assertEqual(data.get('charges_personnel'), 350000)
        self.assertEqual(data.get('excedent_brut'), 435000)
    
    def test_load_nonexistent_file(self):
        """Test avec un fichier inexistant"""
        data = self.loader.load_excel_template("nonexistent_file.xlsx")
        self.assertIsNone(data)
    
    def test_load_invalid_excel_structure(self):
        """Test avec un fichier Excel de structure invalide"""
        invalid_path = self.create_invalid_excel_file()
        data = self.loader.load_excel_template(invalid_path)
        
        # Le chargeur devrait essayer l'extraction générique
        self.assertIsInstance(data, dict)
        os.remove(invalid_path)
    
    def test_analyze_workbook_structure(self):
        """Test de l'analyse de la structure du classeur"""
        workbook = openpyxl.load_workbook(self.temp_excel_path)
        sheets_info = self.loader.analyze_workbook_structure(workbook)
        
        self.assertIn('sheet_names', sheets_info)
        self.assertIn('has_bilan', sheets_info)
        self.assertIn('has_cr', sheets_info)
        self.assertIn('has_tft', sheets_info)
        
        # Vérifier la détection des feuilles BCEAO
        self.assertTrue(sheets_info['has_bilan'])
        self.assertTrue(sheets_info['has_cr'])
        self.assertTrue(sheets_info['has_tft'])
        
        workbook.close()
    
    def test_is_bceao_template(self):
        """Test de la détection du modèle BCEAO"""
        workbook = openpyxl.load_workbook(self.temp_excel_path)
        sheets_info = self.loader.analyze_workbook_structure(workbook)
        
        is_bceao = self.loader.is_bceao_template(sheets_info)
        self.assertTrue(is_bceao)
        
        workbook.close()
    
    def test_extract_bilan_sheet(self):
        """Test de l'extraction des données du bilan"""
        workbook = openpyxl.load_workbook(self.temp_excel_path)
        bilan_sheet = workbook['Bilan']
        
        bilan_data = self.loader.extract_bilan_sheet(bilan_sheet)
        
        # Vérifier les données extraites
        self.assertEqual(bilan_data.get('terrains'), 200000)
        self.assertEqual(bilan_data.get('batiments'), 500000)
        self.assertEqual(bilan_data.get('capital'), 200000)
        self.assertEqual(bilan_data.get('reserves'), 200000)  # Somme des réserves
        
        workbook.close()
    
    def test_extract_cr_sheet(self):
        """Test de l'extraction du compte de résultat"""
        workbook = openpyxl.load_workbook(self.temp_excel_path)
        cr_sheet = workbook['CR']
        
        cr_data = self.loader.extract_cr_sheet(cr_sheet)
        
        # Vérifier les données extraites
        self.assertEqual(cr_data.get('chiffre_affaires'), 1700000)
        self.assertEqual(cr_data.get('valeur_ajoutee'), 785000)
        self.assertEqual(cr_data.get('charges_personnel'), 350000)
        self.assertEqual(cr_data.get('resultat_net'), 150000)
        
        # Vérifier le calcul des charges d'exploitation
        expected_charges = (120000 + 600000 + 200000 + 50000 + 100000 + 
                          30000 + 40000 + 350000 + 100000)
        self.assertEqual(cr_data.get('charges_exploitation'), expected_charges)
        
        workbook.close()
    
    def test_extract_tft_sheet(self):
        """Test de l'extraction du tableau des flux"""
        workbook = openpyxl.load_workbook(self.temp_excel_path)
        tft_sheet = workbook['TFT']
        
        tft_data = self.loader.extract_tft_sheet(tft_sheet)
        
        # Vérifier les données extraites
        self.assertEqual(tft_data.get('tresorerie_ouverture'), 30000)
        self.assertEqual(tft_data.get('cafg'), 250000)
        self.assertEqual(tft_data.get('flux_activites_operationnelles'), 200000)
        self.assertEqual(tft_data.get('tresorerie_cloture'), 80000)
        
        workbook.close()
    
    def test_get_cell_value(self):
        """Test de l'extraction de valeurs de cellules"""
        workbook = openpyxl.load_workbook(self.temp_excel_path)
        sheet = workbook['Bilan']
        
        # Test valeur numérique normale
        value = self.loader.get_cell_value(sheet, 'E11')
        self.assertEqual(value, 200000.0)
        
        # Test cellule vide
        value = self.loader.get_cell_value(sheet, 'Z99')
        self.assertEqual(value, 0.0)
        
        # Test avec une cellule contenant du texte
        sheet['A1'] = "Text value"
        value = self.loader.get_cell_value(sheet, 'A1')
        self.assertEqual(value, 0.0)
        
        # Test avec une chaîne numérique
        sheet['A2'] = "123,456.78"
        value = self.loader.get_cell_value(sheet, 'A2')
        self.assertEqual(value, 123456.78)
        
        # Test avec valeur négative entre parenthèses
        sheet['A3'] = "(50000)"
        value = self.loader.get_cell_value(sheet, 'A3')
        self.assertEqual(value, -50000.0)
        
        workbook.close()
    
    def test_find_sheet_by_pattern(self):
        """Test de la recherche de feuilles par pattern"""
        workbook = openpyxl.load_workbook(self.temp_excel_path)
        
        # Test recherche existante
        sheet_name = self.loader.find_sheet_by_pattern(workbook, ['CR', 'Compte'])
        self.assertEqual(sheet_name, 'CR')
        
        # Test recherche inexistante
        sheet_name = self.loader.find_sheet_by_pattern(workbook, ['Inexistant'])
        self.assertIsNone(sheet_name)
        
        workbook.close()
    
    def test_validate_and_clean_data(self):
        """Test de la validation et nettoyage des données"""
        raw_data = {
            'total_actif': 1000000,
            'capitaux_propres': 400000,
            'chiffre_affaires': 1500000,
            'invalid_field': float('nan'),  # Valeur NaN
            'negative_field': -50000,
            'zero_field': 0,
            'text_field': 'invalid'  # Sera converti en 0
        }
        
        cleaned_data = self.loader.validate_and_clean_data(raw_data)
        
        # Vérifier que les valeurs valides sont conservées
        self.assertEqual(cleaned_data['total_actif'], 1000000.0)
        self.assertEqual(cleaned_data['capitaux_propres'], 400000.0)
        self.assertEqual(cleaned_data['negative_field'], -50000.0)
        
        # Vérifier que les valeurs invalides sont nettoyées
        self.assertEqual(cleaned_data['invalid_field'], 0.0)
        
        # Vérifier que les champs manquants sont ajoutés
        self.assertIn('resultat_net', cleaned_data)
        self.assertIn('immobilisations_nettes', cleaned_data)
    
    def test_perform_basic_consistency_checks(self):
        """Test des vérifications de cohérence de base"""
        # Données cohérentes
        consistent_data = {
            'total_actif': 1000000,
            'capitaux_propres': 400000,
            'dettes_financieres': 300000,
            'dettes_court_terme': 300000,
            'chiffre_affaires': 1500000,
            'resultat_net': 75000
        }
        
        # Le test ne devrait pas lever d'exception
        self.loader.perform_basic_consistency_checks(consistent_data)
        
        # Données incohérentes
        inconsistent_data = {
            'total_actif': 1000000,
            'capitaux_propres': 400000,
            'dettes_financieres': 300000,
            'dettes_court_terme': 500000,  # Total passif > actif
            'chiffre_affaires': -1500000,  # CA négatif
            'resultat_net': 75000
        }
        
        # Test avec logging activé pour capturer les warnings
        with self.assertLogs(self.loader.logger, level='WARNING') as log:
            self.loader.perform_basic_consistency_checks(inconsistent_data)
            # Vérifier qu'au moins un warning a été émis
            self.assertGreater(len(log.output), 0)
    
    def test_create_sample_data(self):
        """Test de la création de données d'exemple"""
        sample_data = self.loader.create_sample_data()
        
        # Vérifier que c'est un dictionnaire non vide
        self.assertIsInstance(sample_data, dict)
        self.assertGreater(len(sample_data), 0)
        
        # Vérifier quelques champs obligatoires
        required_fields = ['total_actif', 'capitaux_propres', 'chiffre_affaires', 'resultat_net']
        for field in required_fields:
            self.assertIn(field, sample_data)
            self.assertGreater(sample_data[field], 0)
        
        # Vérifier la cohérence du bilan dans les données d'exemple
        total_passif = (sample_data['capitaux_propres'] + 
                       sample_data['dettes_financieres'] + 
                       sample_data['dettes_court_terme'])
        
        # Tolérance pour l'équilibre du bilan
        ecart = abs(sample_data['total_actif'] - total_passif)
        self.assertLess(ecart, sample_data['total_actif'] * 0.01)  # Moins de 1% d'écart
    
    def test_export_mapping_template(self):
        """Test de l'export du modèle de mapping"""
        mapping_path = os.path.join(self.temp_dir, "mapping_template.json")
        
        # Exporter le template
        self.loader.export_mapping_template(mapping_path)
        
        # Vérifier que le fichier a été créé
        self.assertTrue(os.path.exists(mapping_path))
        
        # Vérifier le contenu
        import json
        with open(mapping_path, 'r', encoding='utf-8') as f:
            mapping_data = json.load(f)
        
        self.assertIn('description', mapping_data)
        self.assertIn('bilan_actif', mapping_data)
        self.assertIn('bilan_passif', mapping_data)
        self.assertIn('compte_resultat', mapping_data)
        
        # Nettoyer
        os.remove(mapping_path)
    
    def test_identify_field_by_pattern(self):
        """Test de l'identification de champs par pattern"""
        # Test avec des patterns existants
        field = self.loader.identify_field_by_pattern("terrains")
        self.assertEqual(field, "terrains")
        
        field = self.loader.identify_field_by_pattern("capital social")
        self.assertEqual(field, "capital")
        
        field = self.loader.identify_field_by_pattern("chiffre d'affaires")
        self.assertEqual(field, "chiffre_affaires")
        
        # Test avec pattern inexistant
        field = self.loader.identify_field_by_pattern("unknown field")
        self.assertIsNone(field)
    
    def test_find_adjacent_value(self):
        """Test de la recherche de valeurs dans les cellules adjacentes"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        # Placer des valeurs test
        sheet['A1'] = "Label"
        sheet['B1'] = 12345  # Valeur à droite
        sheet['A2'] = 67890  # Valeur en dessous
        
        # Test recherche à droite
        value = self.loader.find_adjacent_value(sheet, 1, 1)  # Position A1
        self.assertEqual(value, 12345.0)
        
        # Test recherche en dessous quand pas de valeur à droite
        sheet['B1'] = None
        value = self.loader.find_adjacent_value(sheet, 1, 1)
        self.assertEqual(value, 67890.0)
        
        workbook.close()


class TestExcelLoaderEdgeCases(unittest.TestCase):
    """Tests pour les cas limites et situations d'erreur"""
    
    def setUp(self):
        self.loader = ExcelDataLoader()
        self.temp_dir = tempfile.mkdtemp()
    
    def tearDown(self):
        # Nettoyer les fichiers temporaires
        import shutil
        shutil.rmtree(self.temp_dir)
    
    def test_corrupted_excel_file(self):
        """Test avec un fichier Excel corrompu"""
        corrupted_path = os.path.join(self.temp_dir, "corrupted.xlsx")
        
        # Créer un faux fichier Excel (juste du texte)
        with open(corrupted_path, 'w') as f:
            f.write("This is not an Excel file")
        
        data = self.loader.load_excel_template(corrupted_path)
        self.assertIsNone(data)
        
        os.remove(corrupted_path)
    
    def test_excel_with_formulas(self):
        """Test avec un fichier Excel contenant des formules"""
        formula_path = os.path.join(self.temp_dir, "formulas.xlsx")
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Bilan"
        
        # Ajouter des formules
        sheet['E11'] = 100000  # Valeur de base
        sheet['E12'] = 200000  # Valeur de base
        sheet['E21'] = '=E11+E12'  # Formule
        
        workbook.save(formula_path)
        workbook.close()
        
        # Test avec data_only=True (formules évaluées)
        data = self.loader.load_excel_template(formula_path)
        self.assertIsNotNone(data)
        
        os.remove(formula_path)
    
    def test_excel_with_empty_sheets(self):
        """Test avec des feuilles vides"""
        empty_path = os.path.join(self.temp_dir, "empty.xlsx")
        
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        
        # Créer des feuilles vides avec les bons noms
        bilan_sheet = workbook.create_sheet("Bilan")
        cr_sheet = workbook.create_sheet("CR")
        tft_sheet = workbook.create_sheet("TFT")
        
        workbook.save(empty_path)
        workbook.close()
        
        data = self.loader.load_excel_template(empty_path)
        
        # Devrait retourner des données même si vides
        self.assertIsInstance(data, dict)
        
        os.remove(empty_path)
    
    def test_excel_with_mixed_data_types(self):
        """Test avec différents types de données dans les cellules"""
        mixed_path = os.path.join(self.temp_dir, "mixed_types.xlsx")
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Bilan"
        
        # Différents types de données
        sheet['E11'] = 123456      # Entier
        sheet['E12'] = 123.456     # Float
        sheet['E13'] = "123,456"   # String numérique avec virgule
        sheet['E14'] = "N/A"       # String non numérique
        sheet['E15'] = None        # Valeur nulle
        sheet['E16'] = True        # Booléen
        
        workbook.save(mixed_path)
        workbook.close()
        
        data = self.loader.load_excel_template(mixed_path)
        self.assertIsNotNone(data)
        
        os.remove(mixed_path)
    
    @patch('openpyxl.load_workbook')
    def test_memory_error_handling(self, mock_load):
        """Test de la gestion des erreurs mémoire"""
        mock_load.side_effect = MemoryError("Not enough memory")
        
        data = self.loader.load_excel_template("dummy_path.xlsx")
        self.assertIsNone(data)
    
    @patch('openpyxl.load_workbook')
    def test_permission_error_handling(self, mock_load):
        """Test de la gestion des erreurs de permission"""
        mock_load.side_effect = PermissionError("Access denied")
        
        data = self.loader.load_excel_template("dummy_path.xlsx")
        self.assertIsNone(data)
    
    def test_very_large_numbers(self):
        """Test avec des nombres très grands"""
        large_numbers_path = os.path.join(self.temp_dir, "large_numbers.xlsx")
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Bilan"
        
        # Très grands nombres
        sheet['E11'] = 1e15  # 1 quadrillion
        sheet['E12'] = 1e-6  # Très petit nombre
        sheet['E13'] = float('inf')  # Infini
        
        workbook.save(large_numbers_path)
        workbook.close()
        
        data = self.loader.load_excel_template(large_numbers_path)
        
        # Vérifier que les nombres anormaux sont gérés
        cleaned_data = self.loader.validate_and_clean_data(data)
        self.assertIsInstance(cleaned_data, dict)
        
        os.remove(large_numbers_path)
    
    def test_extract_generic_template_fallback(self):
        """Test du fallback vers l'extraction générique"""
        generic_path = os.path.join(self.temp_dir, "generic.xlsx")
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "FinancialData"  # Nom non standard
        
        # Ajouter des données avec patterns reconnaissables
        sheet['A1'] = "Terrains"
        sheet['B1'] = 150000
        sheet['A2'] = "Capital Social"
        sheet['B2'] = 200000
        sheet['A3'] = "Chiffre d'Affaires"
        sheet['B3'] = 1500000
        
        workbook.save(generic_path)
        workbook.close()
        
        data = self.loader.load_excel_template(generic_path)
        
        # Devrait utiliser l'extraction générique
        self.assertIsInstance(data, dict)
        
        os.remove(generic_path)


class TestExcelLoaderPerformance(unittest.TestCase):
    """Tests de performance pour le chargeur Excel"""
    
    def setUp(self):
        self.loader = ExcelDataLoader()
        self.temp_dir = tempfile.mkdtemp()
    
    def tearDown(self):
        import shutil
        shutil.rmtree(self.temp_dir)
    
    def test_large_excel_file_performance(self):
        """Test de performance avec un gros fichier Excel"""
        large_path = os.path.join(self.temp_dir, "large.xlsx")
        
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        
        # Créer une feuille avec beaucoup de données
        bilan_sheet = workbook.create_sheet("Bilan")
        
        # Remplir 1000 lignes de données
        for i in range(1, 1001):
            bilan_sheet[f'A{i}'] = f"Item {i}"
            bilan_sheet[f'B{i}'] = i * 1000
        
        # Ajouter les cellules importantes pour le test
        bilan_sheet['E11'] = 200000
        bilan_sheet['E35'] = 1000000
        
        workbook.save(large_path)
        workbook.close()
        
        import time
        start_time = time.time()
        
        data = self.loader.load_excel_template(large_path)
        
        end_time = time.time()
        processing_time = end_time - start_time
        
        # Vérifier que le traitement reste raisonnable (< 10 secondes)
        self.assertLess(processing_time, 10.0, "Traitement trop lent pour un gros fichier")
        self.assertIsNotNone(data)
        
        os.remove(large_path)
    
    def test_multiple_sheets_performance(self):
        """Test avec de nombreuses feuilles"""
        multi_sheets_path = os.path.join(self.temp_dir, "multi_sheets.xlsx")
        
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        
        # Créer 20 feuilles
        for i in range(20):
            sheet = workbook.create_sheet(f"Sheet{i}")
            sheet['A1'] = f"Data {i}"
        
        # Ajouter les feuilles importantes
        bilan_sheet = workbook.create_sheet("Bilan")
        bilan_sheet['E35'] = 1000000
        
        cr_sheet = workbook.create_sheet("CR")
        cr_sheet['E46'] = 100000
        
        workbook.save(multi_sheets_path)
        workbook.close()
        
        import time
        start_time = time.time()
        
        data = self.loader.load_excel_template(multi_sheets_path)
        
        end_time = time.time()
        processing_time = end_time - start_time
        
        # Vérifier performance acceptable
        self.assertLess(processing_time, 5.0, "Traitement trop lent avec nombreuses feuilles")
        self.assertIsNotNone(data)
        
        os.remove(multi_sheets_path)


if __name__ == '__main__':
    # Configuration des tests avec niveau de détail élevé
    unittest.main(verbosity=2, buffer=True)
            